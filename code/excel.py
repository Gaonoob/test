
import xlrd
import sys, time
import openpyxl
import os
import re
import json
import copy
#代表有效值
DayValue = ['庚申','庚寅','己未','辛卯','壬辰','壬戌','癸巳','癸亥','甲午','乙未','丙申','丁酉','戊戌','己亥','辛丑','癸卯','壬辰','壬戌','甲辰','乙巳','丙午','丁未','戊申','己酉','庚戌','辛亥','壬子','癸丑','己丑','戊午','戊子']
MonthValue = ['十','十一','十二','一','二','三','四','五','六','七','八','九','后九']
YearValue = ['廿五','廿六','廿七','廿八','廿九','卅','卅一','卅二','卅三','卅四','卅五','卅六','卅七','元']
CharValue = [".","□","〼"," "]
MissingValue = ['一','二','三','四','五','六','七','八','九','后','十','卅','廿'] #当残缺时,使用list匹配
#字符串匹配
YearFlag = ".(?=年)|..(?=年)" #获取年前面的两个或
MonthFlag = ".(?=月)|..(?=月)" #获取月前面的两个字符
ShuoTwoFlag = "..(?=朔)" #获取朔前面的两个字符
DayTwoFlag = "(?<=朔).." #获取朔后面的两个字符表示日
DayThreeFlag = "(?<=月).." #获取月后面的两个字符表示日
ShuoThreeFlag ="(?<=月)..." #获取月后面三个字
#记录时间List
YearTime = [] #记录完整年时间
YearTime1 = [] #记录需要替换年时间
YearTime2 = [] #记录残缺年时间
MonthTime = [] #记录完整月时间
MonthTime1 = [] #记录需要替换月时间
MonthTime2 = [] #记录残缺月时间
DayTime = [] #记录完整日时间
DayTime1 = [] #记录需要替换日时间
DayTime2 = [] #记录残缺日时间
ShuoTime = [] #记录完整朔时间
ShuoTime1 = [] #记录需要替换朔时间
ShuoTime2 = [] #记录残缺朔时间
RecordDict = {} #记录dict ID +完整时间 + 守丞Name
RepalceFlag = False #匹配FLag
AccuratTime = "" #记录完成时间

ExcelPath = "年表.xlsx"
ExcelIndex = xlrd.open_workbook(ExcelPath)
SheetNameList = ExcelIndex.sheet_names()
#打印指定sheet页所有日期
def ExcelSheet(ExcelIndex, Sheetstr):
    SheetIndex = ExcelIndex.sheet_by_name(Sheetstr)
    for j in range(0, SheetIndex.ncols):
        for i in range(2, SheetIndex.nrows):
            DayValue = SheetIndex.cell(i, j).value
            ShuoValue = SheetIndex.cell(2, j).value #每个月的第一天
            MonthValue = SheetIndex.cell(0, j).value
            if(DayValue != "" ):
                print (str(SheetIndex.name)+',' + str(MonthValue)+',' + str(ShuoValue) + '朔,' + str(DayValue)+'日')

#通过输入年月朔日匹配另一种缺失时间 并且生成完整时间
def GetDataTime(Year, Month, Shuo, Day):
    global RepalceFlag
    global AccuratTime
    for k in range(len(SheetNameList)):
        SheetIndex = ExcelIndex.sheet_by_index(k)
        for j in range(0, SheetIndex.ncols):
            for i in range(2, SheetIndex.nrows):
                DayValue = SheetIndex.cell(i, j).value
                ShuoValue = SheetIndex.cell(2, j).value #每个月的第一天
                MonthValue = SheetIndex.cell(0, j).value
                if(DayValue != "" and Month == MonthValue and Shuo == ShuoValue and  DayValue == Day and Year == ""):
                    RepalceFlag = True
                    AccuratTime = str(SheetIndex.name + MonthValue + ShuoValue + DayValue)
                    return SheetIndex.name
                elif(DayValue != "" and Year == SheetIndex.name and Shuo == ShuoValue and  DayValue == Day and Month == ""):
                    RepalceFlag = True
                    AccuratTime = str(SheetIndex.name + MonthValue + ShuoValue + DayValue)
                    return str(MonthValue)
                elif(DayValue != "" and Year == SheetIndex.name and Month == MonthValue  and Shuo == "" and Day == DayValue):
                    RepalceFlag = True
                    AccuratTime = str(SheetIndex.name + MonthValue + ShuoValue + DayValue)
                    return str(ShuoValue)

def GetOneDataTime(Year, Month, Shuo):
    for k in range(len(SheetNameList)):
        SheetIndex = ExcelIndex.sheet_by_index(k)
        for j in range(0, SheetIndex.ncols):
            for i in range(2, SheetIndex.nrows):
                DayValue = SheetIndex.cell(i, j).value
                ShuoValue = SheetIndex.cell(2, j).value #每个月的第一天
                MonthValue = SheetIndex.cell(0, j).value
                if(DayValue != "" and Year == SheetIndex.name and Shuo == ShuoValue  and Month == ""):
                    return str(MonthValue)
                elif(DayValue != "" and Year == SheetIndex.name and Month == MonthValue  and Shuo == ""):
                    return str(ShuoValue)

def JudgeStringIndex(string, src):
    index_list = []
    for idx, char in enumerate(string):
        if char == src:
            index_list.append(idx)
    return index_list


def ComputeShuo(Year, Month ,Shuo):
    if(Year != "" and Month != "" and Shuo == ""):
        return(GetOneDataTime(Year, Month))
    elif (Year != "" and Month != "" and Shuo == ""):
        return(GetOneDataTime(Year, Shuo))


#每行开始重新初始化状态
def LineInitialState():
    global YearTime
    global YearTime1
    global YearTime2
    global MonthTime
    global MonthTime1
    global MonthTime2
    global DayTime
    global DayTime1
    global DayTime2
    global ShuoTime
    global ShuoTime1
    global ShuoTime2
    global AccuratTime
    YearTime = []
    YearTime1 = []
    YearTime2 = []
    MonthTime = []
    MonthTime1 = []
    MonthTime2 = []
    DayTime = []
    DayTime1 = []
    DayTime2 = []
    ShuoTime = []
    ShuoTime1 = []
    ShuoTime2 = []
    AccuratTime = ""

#去掉LIST中重复信息
def RemoveExtraStrings(list):
    listTmp = []
    for i in list:
        if i not in listTmp:
            listTmp.append(i)
    return listTmp

#打开里耶秦简.txt文件 读取每行内容 记录ID 守城 信息
def OpenTxtFile():
    global RepalceFlag
    global SheetDayValue
    global MonthValue
    global YearValue
    global AccuratTime
    global MonthTime

    FileFd = open(r"E:\Anaconda3\envs\FirstDemo\python\里耶秦简.txt", "r", encoding = "utf-8")
    FileFd1 = open(r"E:\Anaconda3\envs\FirstDemo\python\New.txt", "w+", encoding = "utf-8")
    FileFd2 = open(r"E:\Anaconda3\envs\FirstDemo\python\NewYear.txt", "w+", encoding = "utf-8")
    for line in FileFd.readlines():
        LineInitialState()
        IDIndex = line.find("\\")
        NameIndex = JudgeStringIndex(line, "守丞")
        print(NameIndex)
        IDString = line[0:IDIndex]
        NameLen = line.count("守丞")
       # print(line[NameIndex+2:NameIndex+3])
        '''#查找可用年份
        YearList = re.findall(YearFlag, line)
        for i in YearList:
            if i in YearValue:
                YearTime.append(i + "年")
            else:
                if (i[1:2] in YearValue):
                    YearTime.append(i[1:2] + "年") 
                elif (i[1:2] in CharValue): #年前面第1位为特殊字符
                    YearTime1.append(i[1:2] + "年") 
                elif(i[0:1] in MissingValue or i[0:1] in CharValue): #年前面第2位为特殊字符或者有效汉字
                    YearTime2.append(i[0:2] + "年")
        #查找可用月份
        MonthList = re.findall(MonthFlag, line)
        for i in MonthList:
            if i in MonthValue:
                MonthTime.append(i + "月")
            else:
                if (i[1:2] in MonthValue):
                    MonthTime.append(i[1:2] + "月")
                elif (i[1:2] in CharValue):#年前面第1位为特殊字符
                    MonthTime1.append(i[1:2] + "月")
                elif(i[0:1] in MissingValue or i[0:1] in CharValue): #年前面第2位为特殊字符或者有效汉字
                    MonthTime2.append(i[0:2] + "月")
            YearTime = RemoveExtraStrings(YearTime)
            YearTime1 = RemoveExtraStrings(YearTime1)
            YearTime2 = RemoveExtraStrings(YearTime2)
            MonthTime = RemoveExtraStrings(MonthTime)
            MonthTime = RemoveExtraStrings(MonthTime1)
            MonthTime = RemoveExtraStrings(MonthTime2)'''
'''

            for k in MonthTime:
                FileFd1.write(k + IDString + "\n")
            for k in MonthTime1:
                FileFd1.write(i + IDString + "\r")
            for k in MonthTime2:
                FileFd1.write(i + IDString + "\r")
'''

def CheckDataTimeIndex(Year, Month, Shuo, Day):
    for k in range(len(SheetNameList)):
        SheetIndex = ExcelIndex.sheet_by_index(k)
        for j in range(0, SheetIndex.ncols):
            for i in range(2, SheetIndex.nrows):
                DayValue = SheetIndex.cell(i, j).value
                ShuoValue = SheetIndex.cell(2, j).value #每个月的第一天
                MonthValue = SheetIndex.cell(0, j).value
                if(Year == SheetIndex.name and MonthValue == Month and Shuo == ShuoValue and DayValue == Day):
                    return k+j+i
'''
def giveAllTime(string):
    YearString = re.findall(YearTwoFlag, string)
    YearString = str(YearString)
    len1 = len(YearString)
    if(len1 == 7): 
        YearString = YearString[2:5];
    elif len1 == 6: 
        YearString = YearString[2:4];
    
    MonthString = re.findall(MonthFlag, string)
    MonthString = str(MonthString)
    len1 = len(MonthString)
    if(len1 == 7): #个位
        MonthString = MonthString[3:5];
    elif len1 == 8: #两位
        MonthString = MonthString[3:6];
    
    ShuoString = re.findall(ShuowoFlag, string)
    ShuoString = str(ShuoString)
    len1 = len(ShuoString)
    ShuoString = ShuoString[2:4];
    
    DayString = re.findall(DayThreeFlag, string)
    DayString = str(DayString)
    len1 = len(DayString)
    DayString = DayString[2:4];
    
    return (CheckDataTimeIndex(YearString, MonthString, ShuoString, DayString))
   
   
def compareString(src, dest):
    Index1 = giveAllTime(src)
    Index2 = giveAllTime(dest)
    if(Index1 > Index2):
        print (Index1)
'''


def AddInformation(ID, Time, Name, Index):
    RecordDict.setdefault(Time, []).append(Name)
    RecordDict.setdefault(Time, []).append(ID)
    if Index != "":
        RecordDict.setdefault(Time, []).append(Index)

def RecordNewDict():
    FileFd = open(r"E:\Anaconda3\envs\FirstDemo\python\Record.txt", "w+", encoding = "utf-8")
    str = "{:{ocp}<39}{:{ocp}<11}{:{ocp}<3}".format("时间","名字","标签",ocp = " ")
    FileFd.write(str+"\r")
    for key in RecordDict:
        str = "{:{ocp}<13}{:{ocp}<5}{:{ocp}<5}".format(key,RecordDict[key][0],RecordDict[key][1],ocp = chr(12288))
        FileFd.write(str+"\r")
    FileFd.close()

def RecordInformation():
    AddInformation("5-1", "元年七月庚子朔丁未","仓白",2333)
    AddInformation("500-195", "十元年是七月庚子朔壬寅","仓",333)
    tmp = sorted(RecordDict.items() ,key = lambda item:item[1][2])
    print(tmp)
    print(tmp[0][1])
    RecordDict.clear()
    #for key in tmp:
        #AddInformation(RecordDict[key][1],key,RecordDict[key][0],"")
    
if __name__ == "__main__":
    #OpenTxtFile()
    RecordInformation()
    RecordNewDict()



'''
    str = "5-1\（正）\元年七月庚子朔丁未倉守陽敢言之獄佐辨 平 士吏賀具獄\縣官食盡甲寅謁告過所縣鄉以次續食雨留不能決宿齎\來發傳零陽田能自食當騰期卅日敢言之 七月戊申零陽\襲移過所縣鄉 静手 七月庚子朔癸亥遷陵守丞固告倉啬夫\以律令從事 嘉手\（背）\遷陵食辨平書己巳旦□□□□遷陵\七月癸亥旦士五臂以來 嘉發"
    MonthList = re.findall(MonthFlag, str)
    for i in MonthList:
        if i in MonthValue:
            MonthTime.append(i + "月")
        else:
            if (i[1:2] in MonthValue):
                MonthTime.append(i[1:2] + "月")
            elif (i[1:2] in CharValue):#年前面第1位为特殊字符
                MonthTime1.append(i[1:2] + "月")
            elif(i[0:1] in MissingValue or i[0:1] in CharValue): #年前面第2位为特殊字符或者有效汉字
                MonthTime2.append(i[0:2] + "月")
    MonthTime = RemoveExtraStrings(MonthTime)
    print (MonthTime)
'''                

            
                


    


